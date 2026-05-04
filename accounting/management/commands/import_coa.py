from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from accounting.models import Account


class Command(BaseCommand):
    help = "Import chart of accounts from an Excel file with columns: code, Arabic sub account, Sub account, Arabic main account, Main account."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to COA.xlsx")

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("openpyxl is required. Run: pip install -r requirements.txt") from exc

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]

        created = 0
        updated = 0
        skipped = 0

        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            code, name_ar, name_en, main_account_ar, main_account_en = (list(row) + [None] * 5)[:5]
            code = str(code or "").strip()
            name_ar = str(name_ar or "").strip()

            if not code or not name_ar:
                skipped += 1
                continue

            defaults = {
                "name_ar": name_ar,
                "name_en": str(name_en or "").strip(),
                "main_account_ar": str(main_account_ar or "").strip(),
                "main_account_en": str(main_account_en or "").strip(),
                "account_type": Account.infer_type(code),
                "is_active": True,
            }
            _, was_created = Account.objects.update_or_create(code=code, defaults=defaults)
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"COA import complete. Created: {created}, Updated: {updated}, Skipped: {skipped}"
            )
        )
